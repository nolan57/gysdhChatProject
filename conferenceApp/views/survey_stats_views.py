from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator

from ..models.survey import Survey
from ..models.conference import Conference
from ..services.survey_stats_service import SurveyStatsService

@login_required
def user_survey_stats(request):
    """查看用户的问卷完成情况"""
    stats = SurveyStatsService.get_user_survey_stats(request.user)
    
    return render(request, 'conferenceApp/survey/user_survey_stats.html', {
        'stats': stats
    })

@login_required
@user_passes_test(lambda u: u.is_staff)
def survey_completion_stats(request, survey_id):
    """查看问卷的整体完成情况"""
    survey = get_object_or_404(Survey, id=survey_id, created_by=request.user)
    stats = SurveyStatsService.get_survey_completion_stats(survey)
    
    # 获取参与者完成列表
    completion_list = []
    for conference in survey.conferences.all():
        participants = SurveyStatsService.get_participant_completion_list(conference, survey)
        completion_list.extend(participants)
    
    # 分页
    paginator = Paginator(completion_list, 20)
    page = request.GET.get('page')
    participants = paginator.get_page(page)
    
    return render(request, 'conferenceApp/survey/survey_completion_stats.html', {
        'survey': survey,
        'stats': stats,
        'participants': participants
    })

@login_required
@user_passes_test(lambda u: u.is_staff)
def conference_survey_stats(request, conference_id):
    """会议问卷完成度统计页面"""
    conference = get_object_or_404(Conference, id=conference_id)
    stats = SurveyStatsService.get_conference_survey_stats(conference)
    participants = SurveyStatsService.get_participant_completion_list(conference)
    
    context = {
        'conference': conference,
        'stats': stats,
        'participants': participants
    }
    
    return render(request, 'conferenceApp/survey/conference_survey_stats.html', context)

@login_required
@user_passes_test(lambda u: u.is_staff)
def export_completion_stats(request, conference_id):
    """导出完成情况统计"""
    conference = get_object_or_404(Conference, id=conference_id)
    stats = SurveyStatsService.get_conference_survey_stats(conference)
    completion_list = SurveyStatsService.get_participant_completion_list(conference)
    
    # 创建Excel文件
    import pandas as pd
    from io import BytesIO
    
    # 准备数据
    data = []
    for item in completion_list:
        data.append({
            '参与者': item['participant'].user.get_full_name(),
            '邮箱': item['participant'].user.email,
            '需完成问卷数': item['total_required'],
            '已完成问卷数': item['completed_count'],
            '完成率': f"{item['completion_rate']}%"
        })
    
    # 创建DataFrame
    df = pd.DataFrame(data)
    
    # 创建Excel文件
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='完成情况', index=False)
    
    # 返回文件
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{conference.name}_问卷完成情况.xlsx"'
    return response

@login_required
@user_passes_test(lambda u: u.is_staff)
def export_completion_report(request, conference_id):
    """导出完成情况统计报告"""
    conference = get_object_or_404(Conference, id=conference_id)
    report = SurveyStatsService.generate_survey_completion_report(conference)
    
    # 创建Excel文件
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    
    # 创建工作簿
    wb = Workbook()
    
    # 文本报告页
    ws_report = wb.active
    ws_report.title = '统计报告'
    
    # 设置报告样式
    title_font = Font(bold=True, size=14, color='FFFFFF')
    title_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    
    ws_report.cell(row=1, column=1, value='问卷完成度统计报告').font = title_font
    ws_report.cell(row=1, column=1).fill = title_fill
    
    # 写入报告内容
    report_lines = report['text_report'].split('\n')
    for idx, line in enumerate(report_lines, start=2):
        ws_report.cell(row=idx, column=1, value=line)
    
    # 问卷完成情况页
    ws_surveys = wb.create_sheet('问卷完成情况')
    survey_headers = ['问卷名称', '总参与者', '完成人数', '完成率']
    ws_surveys.append(survey_headers)
    
    for label, total, completed, rate in zip(
        report['chart_data']['labels'], 
        report['chart_data']['total_participants'], 
        report['chart_data']['completed_counts'], 
        report['chart_data']['completion_rates']
    ):
        ws_surveys.append([label, total, completed, f"{rate}%"])
    
    # 调整列宽
    for col in ws_surveys.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        ws_surveys.column_dimensions[column].width = max_length + 2
    
    # 保存文件
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 返回文件
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{conference.name}_问卷完成度报告.xlsx"'
    return response

@login_required
@user_passes_test(lambda u: u.is_staff)
def survey_completion_chart(request, conference_id):
    """获取问卷完成情况图表数据"""
    conference = get_object_or_404(Conference, id=conference_id)
    chart_data = SurveyStatsService.get_survey_completion_chart_data(conference)
    
    return JsonResponse(chart_data)
