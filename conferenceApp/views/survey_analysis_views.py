from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse, HttpResponse
from django.db.models import Count, Q
from django.db.models.functions import TruncDate, TruncWeek, TruncMonth
import json
from collections import defaultdict
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill

from ..models.survey import Survey, SurveyResponse

@login_required
@user_passes_test(lambda u: u.is_staff)
def survey_analysis(request, survey_id):
    """问卷分析视图"""
    survey = get_object_or_404(Survey, id=survey_id, created_by=request.user)
    responses = survey.responses.all()
    
    # 基本统计
    stats = {
        'total_responses': responses.count(),
        'response_by_date': list(
            responses.annotate(date=TruncDate('submitted_at'))
            .values('date')
            .annotate(count=Count('id'))
            .order_by('date')
        ),
        'response_by_conference': list(
            responses.values('conference__name')
            .annotate(count=Count('id'))
            .order_by('-count')
        )
    }
    
    # 问题回答分析
    questions_analysis = analyze_responses(responses, survey.formio_schema)
    
    return render(request, 'conferenceApp/survey/analysis.html', {
        'survey': survey,
        'stats': stats,
        'questions_analysis': questions_analysis
    })

@login_required
@user_passes_test(lambda u: u.is_staff)
def survey_analysis_data(request, survey_id):
    """获取问卷分析数据的API"""
    survey = get_object_or_404(Survey, id=survey_id, created_by=request.user)
    responses = survey.responses.all()
    
    analysis_type = request.GET.get('type', 'basic')
    
    if analysis_type == 'basic':
        data = {
            'total_responses': responses.count(),
            'response_by_date': list(
                responses.annotate(date=TruncDate('submitted_at'))
                .values('date')
                .annotate(count=Count('id'))
                .order_by('date')
            ),
            'response_by_conference': list(
                responses.values('conference__name')
                .annotate(count=Count('id'))
                .order_by('-count')
            )
        }
    elif analysis_type == 'questions':
        data = analyze_responses(responses, survey.formio_schema)
    else:
        data = {}
    
    return JsonResponse(data)

@login_required
@user_passes_test(lambda u: u.is_staff)
def survey_cross_analysis(request, survey_id):
    """问卷交叉分析视图"""
    survey = get_object_or_404(Survey, id=survey_id, created_by=request.user)
    responses = survey.responses.all()
    
    # 获取所有可用于交叉分析的问题
    components = get_form_components(survey.formio_schema)
    analyzable_components = {
        key: comp for key, comp in components.items()
        if comp.get('type') in ['select', 'radio', 'selectboxes']
    }
    
    # 获取用户选择的交叉分析维度
    dimension1 = request.GET.get('dim1')
    dimension2 = request.GET.get('dim2')
    
    cross_analysis = None
    if dimension1 and dimension2 and dimension1 in analyzable_components and dimension2 in analyzable_components:
        cross_analysis = perform_cross_analysis(
            responses, 
            dimension1, 
            dimension2, 
            analyzable_components[dimension1],
            analyzable_components[dimension2]
        )
    
    return render(request, 'conferenceApp/survey/cross_analysis.html', {
        'survey': survey,
        'components': analyzable_components,
        'cross_analysis': cross_analysis,
        'selected_dim1': dimension1,
        'selected_dim2': dimension2
    })

@login_required
@user_passes_test(lambda u: u.is_staff)
def survey_trend_analysis(request, survey_id):
    """问卷趋势分析视图"""
    survey = get_object_or_404(Survey, id=survey_id, created_by=request.user)
    responses = survey.responses.all()
    
    # 获取所有可分析的问题
    components = get_form_components(survey.formio_schema)
    analyzable_components = {
        key: comp for key, comp in components.items()
        if comp.get('type') in ['select', 'radio', 'selectboxes', 'number']
    }
    
    # 获取用户选择的问题和时间范围
    question_key = request.GET.get('question')
    time_range = request.GET.get('range', 'monthly')  # daily, weekly, monthly
    
    trend_data = None
    if question_key and question_key in analyzable_components:
        trend_data = analyze_trend(
            responses,
            question_key,
            analyzable_components[question_key],
            time_range
        )
    
    return render(request, 'conferenceApp/survey/trend_analysis.html', {
        'survey': survey,
        'components': analyzable_components,
        'trend_data': trend_data,
        'selected_question': question_key,
        'selected_range': time_range
    })

@login_required
@user_passes_test(lambda u: u.is_staff)
def export_survey_data(request, survey_id):
    """导出问卷数据"""
    survey = get_object_or_404(Survey, id=survey_id, created_by=request.user)
    responses = survey.responses.all()
    
    # 获取问题信息
    components = get_form_components(survey.formio_schema)
    
    # 准备数据
    data = []
    for response in responses:
        row = {
            '提交时间': response.submitted_at,
            '会议': response.conference.name if response.conference else '',
        }
        
        for key, comp in components.items():
            answer = response.response_data.get(key)
            if isinstance(answer, list):
                answer = ', '.join(map(str, answer))
            row[comp.get('label', key)] = answer
        
        data.append(row)
    
    # 创建DataFrame
    df = pd.DataFrame(data)
    
    # 根据导出格式选择不同的处理方法
    export_format = request.GET.get('format', 'excel')
    
    if export_format == 'excel':
        # 创建Excel文件
        wb = Workbook()
        ws = wb.active
        ws.title = "问卷回复"
        
        # 写入数据
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        # 调整列宽
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # 返回Excel文件
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{survey.title}_数据.xlsx"'
        wb.save(response)
        return response
        
    elif export_format == 'csv':
        # 导出CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{survey.title}_数据.csv"'
        df.to_csv(response, index=False, encoding='utf-8-sig')
        return response
    
    else:
        # 导出JSON
        response = HttpResponse(content_type='application/json')
        response['Content-Disposition'] = f'attachment; filename="{survey.title}_数据.json"'
        df.to_json(response, orient='records', force_ascii=False)
        return response

def analyze_responses(responses, formio_schema):
    """分析问卷回答"""
    analysis = {}
    components = get_form_components(formio_schema)
    
    for response in responses:
        response_data = response.response_data
        for key, component in components.items():
            if key not in analysis:
                analysis[key] = {
                    'question': component.get('label', ''),
                    'type': component.get('type', ''),
                    'answers': defaultdict(int),
                    'chart_types': get_suitable_chart_types(component.get('type', ''))
                }
            
            answer = response_data.get(key)
            if answer:
                if isinstance(answer, list):
                    for item in answer:
                        analysis[key]['answers'][str(item)] += 1
                else:
                    analysis[key]['answers'][str(answer)] += 1
    
    # 计算百分比
    total_responses = responses.count()
    if total_responses > 0:
        for key in analysis:
            total_answers = sum(analysis[key]['answers'].values())
            analysis[key]['percentages'] = {
                answer: (count / total_answers * 100)
                for answer, count in analysis[key]['answers'].items()
            }
    
    return analysis

def perform_cross_analysis(responses, dim1, dim2, comp1, comp2):
    """执行交叉分析"""
    # 初始化交叉分析矩阵
    matrix = defaultdict(lambda: defaultdict(int))
    totals_dim1 = defaultdict(int)
    totals_dim2 = defaultdict(int)
    
    # 获取选项列表
    options1 = get_component_options(comp1)
    options2 = get_component_options(comp2)
    
    # 统计交叉频次
    for response in responses:
        data = response.response_data
        val1 = data.get(dim1)
        val2 = data.get(dim2)
        
        if val1 and val2:
            # 处理多选题
            if isinstance(val1, list):
                val1 = ', '.join(val1)
            if isinstance(val2, list):
                val2 = ', '.join(val2)
            
            matrix[val1][val2] += 1
            totals_dim1[val1] += 1
            totals_dim2[val2] += 1
    
    # 计算百分比
    total_responses = len(responses)
    percentages = {
        'dim1': {val: (count / total_responses * 100) for val, count in totals_dim1.items()},
        'dim2': {val: (count / total_responses * 100) for val, count in totals_dim2.items()},
        'cross': defaultdict(dict)
    }
    
    for val1 in matrix:
        for val2 in matrix[val1]:
            if totals_dim1[val1] > 0:  # 避免除以0
                percentages['cross'][val1][val2] = (matrix[val1][val2] / totals_dim1[val1] * 100)
    
    return {
        'matrix': dict(matrix),
        'totals_dim1': dict(totals_dim1),
        'totals_dim2': dict(totals_dim2),
        'percentages': percentages,
        'dim1_label': comp1.get('label', dim1),
        'dim2_label': comp2.get('label', dim2),
        'options1': options1,
        'options2': options2
    }

def analyze_trend(responses, question_key, component, time_range):
    """分析问题答案的时间趋势"""
    # 按时间范围分组
    if time_range == 'daily':
        responses = responses.annotate(
            period=TruncDate('submitted_at')
        )
    elif time_range == 'weekly':
        responses = responses.annotate(
            period=TruncWeek('submitted_at')
        )
    else:  # monthly
        responses = responses.annotate(
            period=TruncMonth('submitted_at')
        )
    
    # 获取所有可能的选项
    options = get_component_options(component)
    
    # 初始化趋势数据
    trend_data = defaultdict(lambda: defaultdict(int))
    totals = defaultdict(int)
    
    # 统计每个时间段的答案分布
    for response in responses:
        period = response.period
        answer = response.response_data.get(question_key)
        
        if answer:
            # 处理多选题
            if isinstance(answer, list):
                for item in answer:
                    trend_data[period][str(item)] += 1
                    totals[period] += 1
            else:
                trend_data[period][str(answer)] += 1
                totals[period] += 1
    
    # 计算百分比
    percentages = defaultdict(dict)
    for period in trend_data:
        if totals[period] > 0:  # 避免除以0
            for answer in trend_data[period]:
                percentages[period][answer] = (
                    trend_data[period][answer] / totals[period] * 100
                )
    
    # 准备数据系列
    series_data = []
    periods = sorted(trend_data.keys())
    
    for option in options:
        series = {
            'name': str(option),
            'data': [
                {
                    'x': period.isoformat(),
                    'y': percentages[period].get(str(option), 0)
                }
                for period in periods
            ]
        }
        series_data.append(series)
    
    return {
        'periods': [p.isoformat() for p in periods],
        'series': series_data,
        'question_label': component.get('label', question_key),
        'totals': {p.isoformat(): totals[p] for p in periods}
    }

def get_component_options(component):
    """获取组件的选项列表"""
    if component['type'] == 'select':
        return [item['value'] for item in component.get('data', {}).get('values', [])]
    elif component['type'] == 'radio':
        return [item['value'] for item in component.get('values', [])]
    elif component['type'] == 'selectboxes':
        return list(component.get('values', {}).keys())
    return []

def get_form_components(schema):
    """从Form.io schema中提取组件"""
    components = {}
    
    def extract_components(items):
        for item in items:
            if item.get('key') and item.get('type') not in ['button', 'container', 'panel']:
                components[item['key']] = item
            if 'components' in item:
                extract_components(item['components'])
    
    if 'components' in schema:
        extract_components(schema['components'])
    
    return components

def get_suitable_chart_types(component_type):
    """根据问题类型返回适合的图表类型"""
    if component_type in ['select', 'radio']:
        return ['pie', 'bar', 'doughnut', 'polarArea']
    elif component_type in ['selectboxes', 'checkbox']:
        return ['bar', 'horizontalBar', 'radar']
    elif component_type == 'number':
        return ['bar', 'line', 'scatter']
    else:
        return ['bar']  # 默认使用柱状图
